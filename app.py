"""
Flask 全端使用者登入系統
後端: Flask + SQLite3 + Jinja2
功能: 登入/註冊/登出/信箱驗證/忘記密碼/重設密碼/首頁/修改個人資料/Token
"""
import io
import os
import sqlite3
import hashlib
import secrets
import jwt
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from datetime import datetime, timedelta
from pathlib import Path
from functools import wraps
from flask import Flask, render_template, g, request, redirect, url_for, session, flash, jsonify, send_file

app = Flask(__name__)
app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", "your-secret-key-change-in-production")
app.config["JWT_SECRET_KEY"] = os.environ.get("JWT_SECRET_KEY", "jwt-secret-key-change-in-production")
app.config["JWT_ALGORITHM"] = "HS256"
app.config["JWT_EXPIRATION_DELTA"] = timedelta(hours=24)

# 郵件設定（可透過環境變數設定）
app.config["MAIL_SERVER"] = os.environ.get("MAIL_SERVER", "smtp.gmail.com")
app.config["MAIL_PORT"] = int(os.environ.get("MAIL_PORT", "587"))
app.config["MAIL_USE_TLS"] = os.environ.get("MAIL_USE_TLS", "True").lower() == "true"
app.config["MAIL_USERNAME"] = os.environ.get("MAIL_USERNAME", "")
app.config["MAIL_PASSWORD"] = os.environ.get("MAIL_PASSWORD", "")
app.config["MAIL_FROM"] = os.environ.get("MAIL_FROM", app.config["MAIL_USERNAME"])

# SQLite3 資料庫路徑（Vercel 上未設定時自動使用 /tmp/app.db，避免寫入唯讀檔案系統）
if os.environ.get("DATABASE_PATH"):
    DATABASE = Path(os.environ["DATABASE_PATH"])
elif os.environ.get("VERCEL"):
    DATABASE = Path("/tmp/app.db")
else:
    DATABASE = Path(__file__).parent / "instance" / "app.db"

# 個人資料選項（工作轄區、身分）
WORK_REGION_CHOICES = ["", "北北基", "桃竹苗", "中彰投", "雲嘉南", "高屏"]
ROLE_CHOICES = ["一般使用者", "管理者"]
DEFAULT_ROLE = "一般使用者"


def get_db():
    """取得 SQLite 連線"""
    if "db" not in g:
        try:
            DATABASE.parent.mkdir(parents=True, exist_ok=True)
        except OSError:
            pass  # 例如 /tmp 已存在或唯讀環境
        g.db = sqlite3.connect(str(DATABASE))
        g.db.row_factory = sqlite3.Row
    return g.db


@app.teardown_appcontext
def close_db(exception=None):
    """關閉資料庫連線"""
    db = g.pop("db", None)
    if db is not None:
        db.close()


@app.before_request
def ensure_session_role():
    """若已登入但 session 沒有 role（例如舊登入），從資料庫補上，供側邊欄判斷是否顯示資料庫管理"""
    if "user_id" in session and "role" not in session:
        try:
            db = get_db()
            row = db.execute("SELECT role FROM users WHERE id = ?", (session["user_id"],)).fetchone()
            if row:
                session["role"] = row["role"] or DEFAULT_ROLE
        except Exception:
            pass


def init_db():
    """初始化資料庫"""
    db = get_db()
    cursor = db.cursor()
    
    # 使用者表
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            email TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            email_verified INTEGER DEFAULT 0,
            verification_token TEXT,
            reset_token TEXT,
            reset_token_expires DATETIME,
            birthday DATE,
            phone TEXT,
            address TEXT,
            work_region TEXT,
            role TEXT DEFAULT '一般使用者',
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
        )
    """)
    
    # 既有資料庫補加新欄位（遷移）
    for col_sql in [
        "ALTER TABLE users ADD COLUMN birthday DATE",
        "ALTER TABLE users ADD COLUMN phone TEXT",
        "ALTER TABLE users ADD COLUMN address TEXT",
        "ALTER TABLE users ADD COLUMN work_region TEXT",
        "ALTER TABLE users ADD COLUMN role TEXT DEFAULT '一般使用者'",
    ]:
        try:
            cursor.execute(col_sql)
        except sqlite3.OperationalError:
            pass  # 欄位已存在
    
    # Token 表（用於 API 認證）
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS tokens (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            token TEXT UNIQUE NOT NULL,
            expires_at DATETIME NOT NULL,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
        )
    """)
    
    db.commit()


def hash_password(password):
    """密碼雜湊"""
    return hashlib.sha256(password.encode()).hexdigest()


def generate_token(length=32):
    """產生隨機 token"""
    return secrets.token_urlsafe(length)


def generate_jwt_token(user_id, username):
    """產生 JWT token"""
    payload = {
        "user_id": user_id,
        "username": username,
        "exp": datetime.utcnow() + app.config["JWT_EXPIRATION_DELTA"],
        "iat": datetime.utcnow()
    }
    return jwt.encode(payload, app.config["JWT_SECRET_KEY"], algorithm=app.config["JWT_ALGORITHM"])


def verify_jwt_token(token):
    """驗證 JWT token"""
    try:
        payload = jwt.decode(token, app.config["JWT_SECRET_KEY"], algorithms=[app.config["JWT_ALGORITHM"]])
        return payload
    except jwt.ExpiredSignatureError:
        return None
    except jwt.InvalidTokenError:
        return None


def login_required(f):
    """登入驗證裝飾器"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if "user_id" not in session:
            flash("請先登入", "warning")
            return redirect(url_for("index"))
        return f(*args, **kwargs)
    return decorated_function


def admin_required(f):
    """管理者驗證裝飾器（須為登入且 role 為管理者）"""
    @wraps(f)
    @login_required
    def decorated_function(*args, **kwargs):
        db = get_db()
        row = db.execute("SELECT role FROM users WHERE id = ?", (session["user_id"],)).fetchone()
        if not row or row["role"] != "管理者":
            flash("僅管理者可存取此功能", "error")
            return redirect(url_for("home"))
        return f(*args, **kwargs)
    return decorated_function


def email_verified_required(f):
    """信箱驗證裝飾器"""
    @wraps(f)
    @login_required
    def decorated_function(*args, **kwargs):
        db = get_db()
        user = db.execute(
            "SELECT email_verified FROM users WHERE id = ?", (session["user_id"],)
        ).fetchone()
        if not user or not user["email_verified"]:
            flash("請先驗證您的電子信箱", "warning")
            return redirect(url_for("verify_email"))
        return f(*args, **kwargs)
    return decorated_function


def send_email(to_email, subject, html_body, text_body=None):
    """發送電子郵件"""
    # 如果沒有設定郵件帳號，則不發送（開發環境）
    if not app.config["MAIL_USERNAME"] or not app.config["MAIL_PASSWORD"]:
        print(f"[開發模式] 郵件不會實際發送")
        print(f"收件人: {to_email}")
        print(f"主旨: {subject}")
        print(f"內容: {text_body or html_body}")
        return True
    
    try:
        # 建立郵件
        msg = MIMEMultipart("alternative")
        msg["Subject"] = subject
        msg["From"] = app.config["MAIL_FROM"]
        msg["To"] = to_email
        
        # 添加文字和 HTML 內容
        if text_body:
            part1 = MIMEText(text_body, "plain", "utf-8")
            msg.attach(part1)
        
        part2 = MIMEText(html_body, "html", "utf-8")
        msg.attach(part2)
        
        # 發送郵件
        with smtplib.SMTP(app.config["MAIL_SERVER"], app.config["MAIL_PORT"]) as server:
            if app.config["MAIL_USE_TLS"]:
                server.starttls()
            server.login(app.config["MAIL_USERNAME"], app.config["MAIL_PASSWORD"])
            server.send_message(msg)
        
        return True
    except Exception as e:
        print(f"發送郵件失敗: {e}")
        return False


def send_verification_email(email, username, verification_token):
    """發送驗證郵件"""
    verification_url = url_for("verify_email", token=verification_token, _external=True)
    
    subject = "請驗證您的電子信箱 - Flask 登入系統"
    
    html_body = f"""
    <html>
    <body style="font-family: 'Noto Sans TC', Arial, sans-serif; line-height: 1.6; color: #333;">
        <div style="max-width: 600px; margin: 0 auto; padding: 20px;">
            <h2 style="color: #007bff;">歡迎加入 Flask 登入系統！</h2>
            <p>親愛的 {username}，</p>
            <p>感謝您註冊我們的服務。請點擊下方的按鈕來驗證您的電子信箱：</p>
            <div style="text-align: center; margin: 30px 0;">
                <a href="{verification_url}" style="background-color: #007bff; color: white; padding: 12px 30px; text-decoration: none; border-radius: 5px; display: inline-block;">驗證電子信箱</a>
            </div>
            <p>或者複製以下連結到瀏覽器：</p>
            <p style="word-break: break-all; color: #666;">{verification_url}</p>
            <p style="color: #999; font-size: 12px; margin-top: 30px;">此連結將在 24 小時後過期。</p>
            <p style="color: #999; font-size: 12px;">如果您沒有註冊此帳號，請忽略此郵件。</p>
        </div>
    </body>
    </html>
    """
    
    text_body = f"""
    歡迎加入 Flask 登入系統！
    
    親愛的 {username}，
    
    感謝您註冊我們的服務。請點擊以下連結來驗證您的電子信箱：
    
    {verification_url}
    
    此連結將在 24 小時後過期。
    
    如果您沒有註冊此帳號，請忽略此郵件。
    """
    
    return send_email(email, subject, html_body, text_body)


def send_password_reset_email(email, username, reset_token):
    """發送重設密碼郵件"""
    reset_url = url_for("reset_password", token=reset_token, _external=True)
    
    subject = "重設您的密碼 - Flask 登入系統"
    
    html_body = f"""
    <html>
    <body style="font-family: 'Noto Sans TC', Arial, sans-serif; line-height: 1.6; color: #333;">
        <div style="max-width: 600px; margin: 0 auto; padding: 20px;">
            <h2 style="color: #dc3545;">重設密碼請求</h2>
            <p>親愛的 {username}，</p>
            <p>我們收到了您重設密碼的請求。請點擊下方的按鈕來重設您的密碼：</p>
            <div style="text-align: center; margin: 30px 0;">
                <a href="{reset_url}" style="background-color: #dc3545; color: white; padding: 12px 30px; text-decoration: none; border-radius: 5px; display: inline-block;">重設密碼</a>
            </div>
            <p>或者複製以下連結到瀏覽器：</p>
            <p style="word-break: break-all; color: #666;">{reset_url}</p>
            <p style="color: #999; font-size: 12px; margin-top: 30px;">此連結將在 1 小時後過期。</p>
            <p style="color: #999; font-size: 12px;">如果您沒有請求重設密碼，請忽略此郵件。您的密碼將不會被更改。</p>
        </div>
    </body>
    </html>
    """
    
    text_body = f"""
    重設密碼請求
    
    親愛的 {username}，
    
    我們收到了您重設密碼的請求。請點擊以下連結來重設您的密碼：
    
    {reset_url}
    
    此連結將在 1 小時後過期。
    
    如果您沒有請求重設密碼，請忽略此郵件。您的密碼將不會被更改。
    """
    
    return send_email(email, subject, html_body, text_body)


# ==================== 路由 ====================

@app.route("/")
def index():
    """首頁：未登入顯示登入頁，已登入導向首頁"""
    if "user_id" in session:
        return redirect(url_for("home"))
    return render_template("login.html")


@app.route("/home")
@login_required
def home():
    """登入後的主畫面（4x5 按鈕）"""
    return render_template("index.html")


@app.route("/option/<int:num>")
@login_required
def option_page(num):
    """選項獨立頁面（num 為 1～20）"""
    if num < 1 or num > 20:
        flash("無此選項", "error")
        return redirect(url_for("home"))
    return render_template("option.html", option_num=num)


# ==================== 資料庫管理（僅管理者） ====================

def _user_columns():
    """users 表欄位（供匯出/匯入）"""
    return [
        "id", "username", "email", "password", "email_verified",
        "birthday", "phone", "address", "work_region", "role",
        "created_at", "updated_at"
    ]


@app.route("/db-manage", methods=["GET", "POST"])
@admin_required
def db_manage():
    """資料庫管理頁：列出 users、新增、刪除"""
    db = get_db()
    if request.method == "POST":
        action = request.form.get("action")
        if action == "add":
            username = request.form.get("username", "").strip()
            email = request.form.get("email", "").strip().lower()
            password = request.form.get("password", "")
            if not username or not email or not password:
                flash("新增時請填寫使用者名稱、電子信箱、密碼", "error")
            elif len(password) < 6:
                flash("密碼至少 6 個字元", "error")
            else:
                try:
                    db.execute(
                        """INSERT INTO users (username, email, password_hash, role)
                           VALUES (?, ?, ?, ?)""",
                        (username, email, hash_password(password), request.form.get("role") or DEFAULT_ROLE)
                    )
                    db.commit()
                    flash("已新增使用者", "success")
                except sqlite3.IntegrityError:
                    flash("使用者名稱或電子信箱已存在", "error")
        elif action == "delete":
            uid = request.form.get("user_id", type=int)
            if uid and uid != session.get("user_id"):
                db.execute("DELETE FROM users WHERE id = ?", (uid,))
                db.commit()
                flash("已刪除該筆資料", "success")
            elif uid == session.get("user_id"):
                flash("無法刪除目前登入者", "error")
            else:
                flash("無效的刪除請求", "error")
        return redirect(url_for("db_manage"))
    users = db.execute(
        """SELECT id, username, email, email_verified, birthday, phone, address, work_region, role, created_at
           FROM users ORDER BY id"""
    ).fetchall()
    return render_template(
        "db_manage.html",
        users=users,
        work_region_choices=WORK_REGION_CHOICES,
        role_choices=ROLE_CHOICES,
    )


@app.route("/db-manage/edit/<int:user_id>", methods=["GET", "POST"])
@admin_required
def db_manage_edit(user_id):
    """編輯單筆使用者"""
    db = get_db()
    user = db.execute(
        "SELECT id, username, email, email_verified, birthday, phone, address, work_region, role FROM users WHERE id = ?",
        (user_id,),
    ).fetchone()
    if not user:
        flash("無此使用者", "error")
        return redirect(url_for("db_manage"))
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        email = request.form.get("email", "").strip().lower()
        password = request.form.get("password", "")
        birthday = request.form.get("birthday", "").strip() or None
        phone = request.form.get("phone", "").strip() or None
        address = request.form.get("address", "").strip() or None
        work_region = request.form.get("work_region", "").strip() or None
        role = request.form.get("role", "").strip() or DEFAULT_ROLE
        if not username or not email:
            flash("使用者名稱與電子信箱為必填", "error")
            return render_template("db_manage_edit.html", user=user, work_region_choices=WORK_REGION_CHOICES, role_choices=ROLE_CHOICES)
        try:
            if password:
                if len(password) < 6:
                    flash("密碼至少 6 個字元", "error")
                    return render_template("db_manage_edit.html", user=user, work_region_choices=WORK_REGION_CHOICES, role_choices=ROLE_CHOICES)
                db.execute(
                    """UPDATE users SET username=?, email=?, password_hash=?, birthday=?, phone=?, address=?, work_region=?, role=?, updated_at=CURRENT_TIMESTAMP WHERE id=?""",
                    (username, email, hash_password(password), birthday, phone, address, work_region, role, user_id),
                )
            else:
                db.execute(
                    """UPDATE users SET username=?, email=?, birthday=?, phone=?, address=?, work_region=?, role=?, updated_at=CURRENT_TIMESTAMP WHERE id=?""",
                    (username, email, birthday, phone, address, work_region, role, user_id),
                )
            db.commit()
            flash("已更新資料", "success")
            return redirect(url_for("db_manage"))
        except sqlite3.IntegrityError:
            flash("使用者名稱或電子信箱已被其他帳號使用", "error")
    return render_template(
        "db_manage_edit.html",
        user=user,
        work_region_choices=WORK_REGION_CHOICES,
        role_choices=ROLE_CHOICES,
    )


@app.route("/db-manage/export")
@admin_required
def db_manage_export():
    """匯出 users 表為 Excel"""
    try:
        from openpyxl import Workbook
    except ImportError:
        flash("請安裝 openpyxl：pip install openpyxl", "error")
        return redirect(url_for("db_manage"))
    db = get_db()
    rows = db.execute(
        """SELECT id, username, email, email_verified, birthday, phone, address, work_region, role, created_at, updated_at
           FROM users ORDER BY id"""
    ).fetchall()
    wb = Workbook()
    ws = wb.active
    ws.title = "users"
    headers = _user_columns()
    ws.append(headers)
    for r in rows:
        ws.append([
            r["id"], r["username"], r["email"], "", r["email_verified"],
            r["birthday"] or "", r["phone"] or "", r["address"] or "", r["work_region"] or "", r["role"] or "",
            r["created_at"] or "", r["updated_at"] or "",
        ])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"users_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
    )


@app.route("/db-manage/import", methods=["POST"])
@admin_required
def db_manage_import():
    """從 Excel 匯入 users（依 id 更新或依 username/email 新增）"""
    if "file" not in request.files:
        flash("請選擇要匯入的 Excel 檔案", "error")
        return redirect(url_for("db_manage"))
    f = request.files["file"]
    if not f or not f.filename or not f.filename.lower().endswith((".xlsx", ".xls")):
        flash("請上傳 .xlsx 或 .xls 檔案", "error")
        return redirect(url_for("db_manage"))
    try:
        from openpyxl import load_workbook
    except ImportError:
        flash("請安裝 openpyxl：pip install openpyxl", "error")
        return redirect(url_for("db_manage"))
    wb = load_workbook(f, read_only=True, data_only=True)
    ws = wb.active
    if not ws:
        flash("Excel 無有效工作表", "error")
        return redirect(url_for("db_manage"))
    def _cell_str(val):
        if val is None:
            return None
        if hasattr(val, "strftime"):
            return val.strftime("%Y-%m-%d")
        s = str(val).strip()
        return s if s else None

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    db = get_db()
    cols = _user_columns()
    added, updated = 0, 0
    for row in rows:
        if not row or len(row) < 3:
            continue
        row = list(row)[: len(cols)]
        while len(row) < len(cols):
            row.append(None)
        try:
            uid = int(row[0]) if row[0] is not None else None
        except (TypeError, ValueError):
            uid = None
        username = (row[1] or "").strip() if row[1] is not None else ""
        email = (row[2] or "").strip().lower() if row[2] is not None else ""
        password_cell = row[3]
        password = (password_cell or "").strip() if password_cell is not None else ""
        email_verified = 1 if row[4] in (1, "1", True, "True") else 0
        birthday = _cell_str(row[5])
        phone = (row[6] or "").strip() or None if row[6] is not None else None
        address = (row[7] or "").strip() or None if row[7] is not None else None
        work_region = (row[8] or "").strip() or None if row[8] is not None else None
        role = (row[9] or "").strip() or DEFAULT_ROLE if row[9] is not None else DEFAULT_ROLE
        if not username or not email:
            continue
        existing = db.execute("SELECT id, password_hash FROM users WHERE id = ?", (uid,)).fetchone() if uid else None
        if existing:
            if password:
                db.execute(
                    """UPDATE users SET username=?, email=?, password_hash=?, email_verified=?, birthday=?, phone=?, address=?, work_region=?, role=?, updated_at=CURRENT_TIMESTAMP WHERE id=?""",
                    (username, email, hash_password(password), email_verified, birthday, phone, address, work_region, role, uid),
                )
            else:
                db.execute(
                    """UPDATE users SET username=?, email=?, email_verified=?, birthday=?, phone=?, address=?, work_region=?, role=?, updated_at=CURRENT_TIMESTAMP WHERE id=?""",
                    (username, email, email_verified, birthday, phone, address, work_region, role, uid),
                )
            updated += 1
        else:
            pwd_hash = hash_password(password) if password else hash_password(secrets.token_urlsafe(8))
            db.execute(
                """INSERT INTO users (username, email, password_hash, email_verified, birthday, phone, address, work_region, role)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (username, email, pwd_hash, email_verified, birthday, phone, address, work_region, role),
            )
            added += 1
    db.commit()
    flash(f"匯入完成：新增 {added} 筆，更新 {updated} 筆", "success")
    return redirect(url_for("db_manage"))


@app.route("/register", methods=["GET", "POST"])
def register():
    """註冊"""
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        email = request.form.get("email", "").strip().lower()
        password = request.form.get("password", "")
        confirm_password = request.form.get("confirm_password", "")
        
        # 驗證
        if not username or not email or not password:
            flash("請填寫所有欄位", "error")
            return render_template("register.html")
        
        if password != confirm_password:
            flash("密碼不一致", "error")
            return render_template("register.html")
        
        if len(password) < 6:
            flash("密碼長度至少 6 個字元", "error")
            return render_template("register.html")
        
        db = get_db()
        
        # 檢查使用者名稱是否已存在
        if db.execute("SELECT id FROM users WHERE username = ?", (username,)).fetchone():
            flash("使用者名稱已存在", "error")
            return render_template("register.html")
        
        # 檢查電子信箱是否已存在
        if db.execute("SELECT id FROM users WHERE email = ?", (email,)).fetchone():
            flash("電子信箱已被註冊", "error")
            return render_template("register.html")
        
        # 建立使用者
        password_hash = hash_password(password)
        verification_token = generate_token()
        
        db.execute(
            """INSERT INTO users (username, email, password_hash, verification_token, role)
               VALUES (?, ?, ?, ?, ?)""",
            (username, email, password_hash, verification_token, DEFAULT_ROLE)
        )
        db.commit()
        
        # 發送驗證郵件
        if send_verification_email(email, username, verification_token):
            flash("註冊成功！我們已發送驗證郵件至您的電子信箱，請查收並點擊連結完成驗證。", "success")
        else:
            verification_url = url_for("verify_email", token=verification_token, _external=True)
            flash(f"註冊成功！但郵件發送失敗，請使用此連結驗證：{verification_url}", "warning")
        
        return redirect(url_for("index"))
    
    return render_template("register.html")


@app.route("/login", methods=["GET", "POST"])
def login():
    """登入（GET 導向首頁，POST 處理登入）"""
    if request.method == "GET":
        if "user_id" in session:
            return redirect(url_for("home"))
        return redirect(url_for("index"))
    
    # POST：處理登入
    username = request.form.get("username", "").strip()
    password = request.form.get("password", "")
    
    if not username or not password:
        flash("請填寫所有欄位", "error")
        return render_template("login.html")
    
    db = get_db()
    password_hash = hash_password(password)
    
    user = db.execute(
        "SELECT id, username, email, email_verified, role FROM users WHERE (username = ? OR email = ?) AND password_hash = ?",
        (username, username, password_hash)
    ).fetchone()
    
    if user:
        session["user_id"] = user["id"]
        session["username"] = user["username"]
        session["role"] = user["role"] or DEFAULT_ROLE
        flash(f"歡迎回來，{user['username']}！", "success")
        return redirect(url_for("home"))
    else:
        flash("使用者名稱或密碼錯誤", "error")
        return render_template("login.html")


@app.route("/logout")
def logout():
    """登出"""
    session.clear()
    flash("已成功登出", "success")
    return redirect(url_for("index"))


@app.route("/verify-email")
def verify_email():
    """信箱驗證頁面"""
    token = request.args.get("token")
    
    if not token:
        if "user_id" not in session:
            flash("請先登入", "warning")
            return redirect(url_for("login"))
        
        db = get_db()
        user = db.execute(
            "SELECT email, email_verified, verification_token FROM users WHERE id = ?",
            (session["user_id"],)
        ).fetchone()
        
        if user and user["email_verified"]:
            flash("您的電子信箱已經驗證過了", "info")
            return redirect(url_for("home"))
        
        return render_template("verify_email.html", user=user)
    
    # 驗證 token
    db = get_db()
    user = db.execute(
        "SELECT id, email_verified FROM users WHERE verification_token = ?",
        (token,)
    ).fetchone()
    
    if not user:
        flash("無效的驗證連結", "error")
        return redirect(url_for("index"))
    
    if user["email_verified"]:
        flash("您的電子信箱已經驗證過了", "info")
        return redirect(url_for("index"))
    
    # 更新驗證狀態
    db.execute(
        "UPDATE users SET email_verified = 1, verification_token = NULL WHERE id = ?",
        (user["id"],)
    )
    db.commit()
    
    flash("電子信箱驗證成功！", "success")
    return redirect(url_for("index"))


@app.route("/resend-verification", methods=["POST"])
@login_required
def resend_verification():
    """重新發送驗證郵件"""
    db = get_db()
    user = db.execute(
        "SELECT email, email_verified, verification_token FROM users WHERE id = ?",
        (session["user_id"],)
    ).fetchone()
    
    if user and user["email_verified"]:
        flash("您的電子信箱已經驗證過了", "info")
        return redirect(url_for("home"))
    
    # 產生新的驗證 token
    verification_token = generate_token()
    db.execute(
        "UPDATE users SET verification_token = ? WHERE id = ?",
        (verification_token, session["user_id"])
    )
    db.commit()
    
    # 發送驗證郵件
    if send_verification_email(user["email"], session.get("username", "使用者"), verification_token):
        flash("驗證郵件已重新發送至您的電子信箱，請查收。", "success")
    else:
        verification_url = url_for("verify_email", token=verification_token, _external=True)
        flash(f"郵件發送失敗，請使用此連結驗證：{verification_url}", "warning")
    
    return redirect(url_for("verify_email"))


@app.route("/forgot-password", methods=["GET", "POST"])
def forgot_password():
    """忘記密碼"""
    if request.method == "POST":
        email = request.form.get("email", "").strip().lower()
        
        if not email:
            flash("請輸入電子信箱", "error")
            return render_template("forgot_password.html")
        
        db = get_db()
        user = db.execute("SELECT id, username FROM users WHERE email = ?", (email,)).fetchone()
        
        if user:
            reset_token = generate_token()
            reset_token_expires = (datetime.utcnow() + timedelta(hours=1)).isoformat()
            
            db.execute(
                "UPDATE users SET reset_token = ?, reset_token_expires = ? WHERE id = ?",
                (reset_token, reset_token_expires, user["id"])
            )
            db.commit()
            
            # 發送重設密碼郵件
            if send_password_reset_email(email, user["username"], reset_token):
                flash("重設密碼連結已發送至您的電子信箱，請查收。", "success")
            else:
                reset_url = url_for("reset_password", token=reset_token, _external=True)
                flash(f"郵件發送失敗，請使用此連結重設密碼：{reset_url}", "warning")
        else:
            # 為了安全，即使使用者不存在也顯示成功訊息
            flash("如果該電子信箱已註冊，重設密碼連結已發送", "success")
        
        return redirect(url_for("index"))
    
    return render_template("forgot_password.html")


@app.route("/reset-password", methods=["GET", "POST"])
def reset_password():
    """重設密碼"""
    token = request.args.get("token")
    
    if not token:
        flash("無效的重設密碼連結", "error")
        return redirect(url_for("forgot_password"))
    
    db = get_db()
    user = db.execute(
        "SELECT id, reset_token_expires FROM users WHERE reset_token = ?",
        (token,)
    ).fetchone()
    
    if not user:
        flash("無效的重設密碼連結", "error")
        return redirect(url_for("forgot_password"))
    
    # 解析日期（SQLite 可能返回字符串）
    try:
        if isinstance(user["reset_token_expires"], str):
            expires = datetime.fromisoformat(user["reset_token_expires"].replace(" ", "T"))
        else:
            expires = user["reset_token_expires"]
        if datetime.utcnow() > expires:
            flash("重設密碼連結已過期", "error")
            return redirect(url_for("forgot_password"))
    except (ValueError, TypeError):
        flash("無效的重設密碼連結", "error")
        return redirect(url_for("forgot_password"))
    
    if request.method == "POST":
        password = request.form.get("password", "")
        confirm_password = request.form.get("confirm_password", "")
        
        if not password or not confirm_password:
            flash("請填寫所有欄位", "error")
            return render_template("reset_password.html", token=token)
        
        if password != confirm_password:
            flash("密碼不一致", "error")
            return render_template("reset_password.html", token=token)
        
        if len(password) < 6:
            flash("密碼長度至少 6 個字元", "error")
            return render_template("reset_password.html", token=token)
        
        password_hash = hash_password(password)
        db.execute(
            "UPDATE users SET password_hash = ?, reset_token = NULL, reset_token_expires = NULL WHERE id = ?",
            (password_hash, user["id"])
        )
        db.commit()
        
        flash("密碼重設成功！請使用新密碼登入", "success")
        return redirect(url_for("index"))
    
    return render_template("reset_password.html", token=token)


@app.route("/profile")
@login_required
def profile():
    """個人資料頁面"""
    db = get_db()
    user = db.execute(
        """SELECT id, username, email, email_verified, created_at,
                  birthday, phone, address, work_region, role FROM users WHERE id = ?""",
        (session["user_id"],)
    ).fetchone()
    
    return render_template("profile.html", user=user)


@app.route("/profile/edit", methods=["GET", "POST"])
@login_required
def edit_profile():
    """修改個人資料"""
    db = get_db()
    
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        email = request.form.get("email", "").strip().lower()
        birthday = request.form.get("birthday", "").strip() or None
        phone = request.form.get("phone", "").strip() or None
        address = request.form.get("address", "").strip() or None
        work_region = request.form.get("work_region", "").strip() or None
        role = request.form.get("role", "").strip() or DEFAULT_ROLE
        current_password = request.form.get("current_password", "")
        new_password = request.form.get("new_password", "")
        confirm_password = request.form.get("confirm_password", "")
        
        user = db.execute(
            "SELECT username, email, password_hash FROM users WHERE id = ?",
            (session["user_id"],)
        ).fetchone()
        
        # 驗證目前密碼（如果要修改密碼）
        if new_password:
            if not current_password:
                flash("請輸入目前密碼", "error")
                return redirect(url_for("edit_profile"))
            
            if hash_password(current_password) != user["password_hash"]:
                flash("目前密碼錯誤", "error")
                return redirect(url_for("edit_profile"))
            
            if new_password != confirm_password:
                flash("新密碼不一致", "error")
                return redirect(url_for("edit_profile"))
            
            if len(new_password) < 6:
                flash("密碼長度至少 6 個字元", "error")
                return redirect(url_for("edit_profile"))
        
        # 檢查使用者名稱是否已被其他人使用
        if username != user["username"]:
            if db.execute("SELECT id FROM users WHERE username = ? AND id != ?", (username, session["user_id"])).fetchone():
                flash("使用者名稱已被使用", "error")
                return redirect(url_for("edit_profile"))
        
        # 檢查電子信箱是否已被其他人使用
        email_changed = False
        if email != user["email"]:
            if db.execute("SELECT id FROM users WHERE email = ? AND id != ?", (email, session["user_id"])).fetchone():
                flash("電子信箱已被使用", "error")
                return redirect(url_for("edit_profile"))
            email_changed = True
        
        # 更新資料
        update_fields = []
        update_values = []
        
        if username != user["username"]:
            update_fields.append("username = ?")
            update_values.append(username)
            session["username"] = username
        
        if email_changed:
            update_fields.append("email = ?")
            update_fields.append("email_verified = 0")
            update_values.append(email)
            verification_token = generate_token()
            update_fields.append("verification_token = ?")
            update_values.append(verification_token)
            
            # 發送新的驗證郵件
            send_verification_email(email, username, verification_token)
        
        if new_password:
            password_hash = hash_password(new_password)
            update_fields.append("password_hash = ?")
            update_values.append(password_hash)
        
        # 個人資料欄位（生日、手機、住址、工作轄區、身分）
        update_fields.extend(["birthday = ?", "phone = ?", "address = ?", "work_region = ?", "role = ?"])
        update_values.extend([birthday, phone, address, work_region, role])
        
        if update_fields:
            update_fields.append("updated_at = CURRENT_TIMESTAMP")
            update_values.append(session["user_id"])
            
            query = f"UPDATE users SET {', '.join(update_fields)} WHERE id = ?"
            db.execute(query, update_values)
            db.commit()
            session["role"] = role
            if email_changed:
                flash("個人資料已更新！請重新驗證您的電子信箱", "success")
            else:
                flash("個人資料已更新", "success")
        else:
            flash("沒有變更", "info")
        
        return redirect(url_for("profile"))
    
    user = db.execute(
        """SELECT id, username, email, email_verified, birthday, phone, address, work_region, role
           FROM users WHERE id = ?""",
        (session["user_id"],)
    ).fetchone()
    
    return render_template(
        "edit_profile.html",
        user=user,
        work_region_choices=WORK_REGION_CHOICES,
        role_choices=ROLE_CHOICES,
    )


# ==================== Token API ====================

@app.route("/api/token", methods=["POST"])
def generate_token_api():
    """產生 API Token (JWT)"""
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        
        if not username or not password:
            return jsonify({"ok": False, "message": "請提供使用者名稱和密碼"}), 400
        
        db = get_db()
        password_hash = hash_password(password)
        
        user = db.execute(
            "SELECT id, username FROM users WHERE (username = ? OR email = ?) AND password_hash = ?",
            (username, username, password_hash)
        ).fetchone()
        
        if not user:
            return jsonify({"ok": False, "message": "使用者名稱或密碼錯誤"}), 401
        
        token = generate_jwt_token(user["id"], user["username"])
        
        return jsonify({
            "ok": True,
            "token": token,
            "expires_in": int(app.config["JWT_EXPIRATION_DELTA"].total_seconds())
        }), 200


@app.route("/api/verify-token", methods=["POST"])
def verify_token_api():
    """驗證 API Token"""
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    
    if not token:
        return jsonify({"ok": False, "message": "請提供 token"}), 400
    
    payload = verify_jwt_token(token)
    
    if not payload:
        return jsonify({"ok": False, "message": "無效或過期的 token"}), 401
    
    return jsonify({
        "ok": True,
        "user_id": payload["user_id"],
        "username": payload["username"]
    }), 200


@app.route("/api/user-info", methods=["GET"])
def user_info_api():
    """取得使用者資訊 (需要 Token)"""
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    
    if not token:
        return jsonify({"ok": False, "message": "請提供 token"}), 400
    
    payload = verify_jwt_token(token)
    
    if not payload:
        return jsonify({"ok": False, "message": "無效或過期的 token"}), 401
    
    db = get_db()
    user = db.execute(
        "SELECT id, username, email, email_verified, created_at FROM users WHERE id = ?",
        (payload["user_id"],)
    ).fetchone()
    
    if not user:
        return jsonify({"ok": False, "message": "使用者不存在"}), 404
    
    return jsonify({
        "ok": True,
        "user": {
            "id": user["id"],
            "username": user["username"],
            "email": user["email"],
            "email_verified": bool(user["email_verified"]),
            "created_at": user["created_at"]
        }
    }), 200


# 模組載入時即初始化資料表（Vercel 等環境不會執行 __main__，須在此執行）
with app.app_context():
    try:
        init_db()
    except Exception as e:
        import sys
        print(f"[init_db] {e}", file=sys.stderr)

if __name__ == "__main__":
    app.run(debug=True)
